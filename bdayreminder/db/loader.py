import os
import datetime

from openpyxl import load_workbook

from bdayreminder.db.base import DBSession
from bdayreminder.db.models import Person

BASE_DIR = os.path.dirname(os.path.realpath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'data/data.xlsx')


def exceldata():
    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook.active

    session = DBSession()
    for row in sheet.iter_rows(row_offset=1):
        person = Person(
            name=row[0].value,
            dob=row[1].value,
            email=row[2].value,
            mobile=row[3].value
        )

        session.add(person)
    session.commit()
    session.close()


def sampledata():
    session = DBSession()
    person1 = Person(
        name='John',
        dob=datetime.datetime.now().date(),
        email='saluasif@gmail.com',
        mobile=9916016104
    )

    session.add(person1)

    person2 = Person(
        name='Jane',
        dob=datetime.datetime.now().date(),
        email='noreplycse@gmail.com',
        mobile=9916016103
    )

    session.add(person2)
    session.commit()
    session.close()
