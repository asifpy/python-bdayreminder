import os
from bdayreminder.db.base import DBSession
from bdayreminder.db.models import Person

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.realpath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'data/data.xlsx')


def load_from_excel():
    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook.active

    session = DBSession()
    for row in sheet.iter_rows(row_offset=1):
        person = Person(
            name=row[0].value,
            dob=row[1].value,
            email=row[2].value,
            mobile=row[3].value
        )

        session.add(person)
    session.commit()
    session.close()


if __name__ == '__main__':
    load_from_excel()
